"""Data export module for Excel and JSON formats."""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from bigvalue_search.config import config
from bigvalue_search.bigvalue_api import SearchResults, BuildingInfo, BusinessInfo, CardSalesInfo
from bigvalue_search.exceptions import ExportError

logger = logging.getLogger(__name__)


def _ensure_output_dir() -> Path:
    """Ensure output directory exists."""
    output_dir = Path(config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _timestamp() -> str:
    """Generate timestamp string for filenames."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ─── Excel Styles ───────────────────────────────────────────────────────────

HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws: Worksheet, num_cols: int) -> None:
    """Apply header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data_rows(ws: Worksheet, num_rows: int, num_cols: int) -> None:
    """Apply data styling to data rows."""
    for row in range(2, num_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_column_width(ws: Worksheet) -> None:
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # Estimate width for CJK characters
                val = str(cell.value)
                length = 0
                for ch in val:
                    if ord(ch) > 127:
                        length += 2
                    else:
                        length += 1
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


# ─── Building Sheet ─────────────────────────────────────────────────────────

BUILDING_HEADERS: list[str] = [
    "건물 ID", "건물명", "지번주소", "도로명주소", "건물유형",
    "연면적(㎡)", "대지면적(㎡)", "층수", "건축년도",
    "위도", "경도", "중심거리(m)",
]


def _write_buildings_sheet(wb: Workbook, buildings: list[BuildingInfo]) -> None:
    """Write buildings data to Excel sheet."""
    ws = wb.active
    ws.title = "건물정보"

    # Headers
    for col, header in enumerate(BUILDING_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    # Data
    for row, b in enumerate(buildings, 2):
        ws.cell(row=row, column=1, value=b.building_id)
        ws.cell(row=row, column=2, value=b.building_name)
        ws.cell(row=row, column=3, value=b.address)
        ws.cell(row=row, column=4, value=b.road_address)
        ws.cell(row=row, column=5, value=b.building_type)
        ws.cell(row=row, column=6, value=b.total_floor_area)
        ws.cell(row=row, column=7, value=b.land_area)
        ws.cell(row=row, column=8, value=b.number_of_floors)
        ws.cell(row=row, column=9, value=b.year_built)
        ws.cell(row=row, column=10, value=b.latitude)
        ws.cell(row=row, column=11, value=b.longitude)
        ws.cell(row=row, column=12, value=b.distance_m)

    _style_header_row(ws, len(BUILDING_HEADERS))
    _style_data_rows(ws, len(buildings), len(BUILDING_HEADERS))
    _auto_column_width(ws)


# ─── Business Sheet ─────────────────────────────────────────────────────────

BUSINESS_HEADERS: list[str] = [
    "사업자 ID", "상호명", "업종", "업태", "대표자명",
    "종업원수", "추정매출", "건물 ID", "건물명",
]


def _write_businesses_sheet(wb: Workbook, businesses: list[BusinessInfo]) -> None:
    """Write businesses data to Excel sheet."""
    ws = wb.create_sheet(title="사업자정보")

    for col, header in enumerate(BUSINESS_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    for row, b in enumerate(businesses, 2):
        ws.cell(row=row, column=1, value=b.business_id)
        ws.cell(row=row, column=2, value=b.business_name)
        ws.cell(row=row, column=3, value=b.business_type)
        ws.cell(row=row, column=4, value=b.business_category)
        ws.cell(row=row, column=5, value=b.representative_name)
        ws.cell(row=row, column=6, value=b.employee_count)
        ws.cell(row=row, column=7, value=b.revenue_estimate)
        ws.cell(row=row, column=8, value=b.building_id)
        ws.cell(row=row, column=9, value=b.building_name)

    _style_header_row(ws, len(BUSINESS_HEADERS))
    _style_data_rows(ws, len(businesses), len(BUSINESS_HEADERS))
    _auto_column_width(ws)


# ─── Card Sales Sheet ───────────────────────────────────────────────────────

CARD_SALES_HEADERS: list[str] = [
    "지역 ID", "지역명", "총매출금액", "점포당평균매출",
    "점포수", "카테고리별매출", "월별매출",
]


def _write_card_sales_sheet(wb: Workbook, sales: list[CardSalesInfo]) -> None:
    """Write card sales data to Excel sheet."""
    ws = wb.create_sheet(title="카드매출정보")

    for col, header in enumerate(CARD_SALES_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    for row, s in enumerate(sales, 2):
        ws.cell(row=row, column=1, value=s.area_id)
        ws.cell(row=row, column=2, value=s.area_name)
        ws.cell(row=row, column=3, value=s.total_sales_amount)
        ws.cell(row=row, column=4, value=s.average_sales_per_store)
        ws.cell(row=row, column=5, value=s.number_of_stores)
        ws.cell(row=row, column=6, value=json.dumps(s.sales_by_category, ensure_ascii=False) if s.sales_by_category else "")
        ws.cell(row=row, column=7, value=json.dumps(s.monthly_sales, ensure_ascii=False) if s.monthly_sales else "")

    _style_header_row(ws, len(CARD_SALES_HEADERS))
    _style_data_rows(ws, len(sales), len(CARD_SALES_HEADERS))
    _auto_column_width(ws)


# ─── Summary Sheet ──────────────────────────────────────────────────────────

def _write_summary_sheet(wb: Workbook, results: SearchResults) -> None:
    """Write summary information sheet."""
    ws = wb.create_sheet(title="검색요약")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50

    summary_data = [
        ("검색일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("중심위도", str(results.search_center.latitude) if results.search_center else "N/A"),
        ("중심경도", str(results.search_center.longitude) if results.search_center else "N/A"),
        ("검색반경(m)", str(results.radius_m)),
        ("발견건물수", str(len(results.buildings))),
        ("발견사업자수", str(len(results.businesses))),
        ("카드매출건수", str(len(results.card_sales))),
    ]

    title_font = Font(name="맑은 고딕", size=14, bold=True, color="2F5496")
    ws.cell(row=1, column=1, value="BigValue 건물 검색 결과").font = title_font
    ws.merge_cells("A1:B1")

    for row, (label, value) in enumerate(summary_data, 3):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)


# ─── Public Export Functions ────────────────────────────────────────────────

def export_to_excel(results: SearchResults, filename: str | None = None) -> Path:
    """
    Export search results to Excel file.

    Args:
        results: SearchResults to export
        filename: Optional custom filename

    Returns:
        Path to the created Excel file.
    """
    output_dir = _ensure_output_dir()
    if filename is None:
        filename = f"bigvalue_search_{_timestamp()}.xlsx"

    filepath = output_dir / filename

    wb = Workbook()
    _write_buildings_sheet(wb, results.buildings)
    _write_businesses_sheet(wb, results.businesses)
    _write_card_sales_sheet(wb, results.card_sales)
    _write_summary_sheet(wb, results)

    wb.save(filepath)
    logger.info("Excel file saved: %s", filepath)
    return filepath


def export_to_json(results: SearchResults, filename: str | None = None) -> Path:
    """
    Export search results to JSON file.

    Args:
        results: SearchResults to export
        filename: Optional custom filename

    Returns:
        Path to the created JSON file.
    """
    output_dir = _ensure_output_dir()
    if filename is None:
        filename = f"bigvalue_search_{_timestamp()}.json"

    filepath = output_dir / filename

    data = {
        "search_info": {
            "timestamp": datetime.now().isoformat(),
            "center": {
                "latitude": results.search_center.latitude if results.search_center else None,
                "longitude": results.search_center.longitude if results.search_center else None,
            },
            "radius_m": results.radius_m,
        },
        "buildings": [
            {
                "building_id": b.building_id,
                "building_name": b.building_name,
                "address": b.address,
                "road_address": b.road_address,
                "building_type": b.building_type,
                "total_floor_area": b.total_floor_area,
                "land_area": b.land_area,
                "number_of_floors": b.number_of_floors,
                "year_built": b.year_built,
                "latitude": b.latitude,
                "longitude": b.longitude,
                "distance_m": b.distance_m,
            }
            for b in results.buildings
        ],
        "businesses": [
            {
                "business_id": b.business_id,
                "business_name": b.business_name,
                "business_type": b.business_type,
                "business_category": b.business_category,
                "representative_name": b.representative_name,
                "employee_count": b.employee_count,
                "revenue_estimate": b.revenue_estimate,
                "building_id": b.building_id,
                "building_name": b.building_name,
            }
            for b in results.businesses
        ],
        "card_sales": [
            {
                "area_id": s.area_id,
                "area_name": s.area_name,
                "total_sales_amount": s.total_sales_amount,
                "average_sales_per_store": s.average_sales_per_store,
                "number_of_stores": s.number_of_stores,
                "sales_by_category": s.sales_by_category,
                "monthly_sales": s.monthly_sales,
            }
            for s in results.card_sales
        ],
        "summary": {
            "total_buildings": len(results.buildings),
            "total_businesses": len(results.businesses),
            "total_card_sales": len(results.card_sales),
        },
    }

    filepath.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("JSON file saved: %s", filepath)
    return filepath


def export_results(
    results: SearchResults,
    output_format: str = "both",
) -> dict[str, Path]:
    """
    Export results in the specified format(s).

    Args:
        results: SearchResults to export
        output_format: "excel", "json", or "both"

    Returns:
        Dict mapping format name to file path.
    """
    output_files: dict[str, Path] = {}

    if output_format in ("excel", "both"):
        output_files["excel"] = export_to_excel(results)

    if output_format in ("json", "both"):
        output_files["json"] = export_to_json(results)

    return output_files
